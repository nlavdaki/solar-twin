r"""Convert the Thissio pyranometer Excel (5 yearly tabs, 15-min) to UTC CSVs and
classify clear-sky days. GPU-free (uv).

Input format (THISSIO-2020-2024_step-15min_FINAL.xlsx): one sheet per year, columns
YEAR, MONTH, DAY, HOUR, MIN, TOTAL AVG (GHI W/m^2), DIFFUSE AVG (W/m^2).

TIME: the station clock is FIXED UTC+2 (Greek standard, NO DST) — established
empirically (clearest-day GHI peak vs pvlib solar noon = +2.0 h in every month,
summer and winter alike). So UTC = clock - 2 h, constant year-round.

CLEAR-SKY: a day is "clear" if measured daytime GHI tracks a pvlib clear-sky
(Ineichen) curve closely (ratio in a band) — empirically clear days ~1.1-1.2,
cloudy days < ~0.7, with a clean gap. The twin renders cloudless scenes, so only
clear-sky days are valid for synthetic-lux vs measured-GHI validation.

Outputs:
  --out-all     : pyranometer_thissio_utc.csv   [timestamp_utc, ghi_measured_wm2, diffuse_wm2]
  --out-clear   : pyranometer_thissio_clearsky_days.csv  [date, clearness_ratio]  (clear days only)

Usage (uv):
  uv run python scripts/read_pyranometer.py \
    --xlsx "…/THISSIO-2020-2024_step-15min_FINAL.xlsx" \
    --lat 37.9719 --lon 23.7186 --alt 100 \
    --out-all "…/data/pyrano/pyranometer_thissio_utc.csv" \
    --out-clear "…/data/pyrano/pyranometer_thissio_clearsky_days.csv"
"""
from __future__ import annotations

import argparse
import os

import numpy as np
import pandas as pd


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True)
    p.add_argument("--lat", type=float, required=True)
    p.add_argument("--lon", type=float, required=True)
    p.add_argument("--alt", type=float, default=100.0)
    p.add_argument("--clock-offset-h", type=float, default=2.0,
                   help="station clock minus UTC (Thissio = +2 fixed, no DST)")
    p.add_argument("--clear-min", type=float, default=0.85,
                   help="min meas/clearsky daytime ratio to call a day clear")
    p.add_argument("--clear-max", type=float, default=1.35,
                   help="max ratio (guards against sensor/model anomalies)")
    p.add_argument("--out-all", required=True)
    p.add_argument("--out-clear", required=True)
    args = p.parse_args()

    import openpyxl
    import pvlib

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    rows = []
    for sn in wb.sheetnames:
        for r in wb[sn].iter_rows(min_row=2, values_only=True):
            if r[0] is None:
                continue
            rows.append(r[:7])
    df = pd.DataFrame(rows, columns=["Y", "M", "D", "h", "mi", "ghi", "dif"])
    df["clock"] = pd.to_datetime(dict(year=df.Y, month=df.M, day=df.D, hour=df.h, minute=df.mi))
    df["timestamp_utc"] = (df["clock"] - pd.Timedelta(hours=args.clock_offset_h)).dt.tz_localize("UTC")
    df["ghi"] = pd.to_numeric(df["ghi"], errors="coerce")
    df["dif"] = pd.to_numeric(df["dif"], errors="coerce")

    # write the full UTC time series
    os.makedirs(os.path.dirname(args.out_all), exist_ok=True)
    out = df[["timestamp_utc", "ghi", "dif"]].copy()
    out.columns = ["timestamp_utc", "ghi_measured_wm2", "diffuse_wm2"]
    out["timestamp_utc"] = out["timestamp_utc"].map(lambda t: t.isoformat())
    out.to_csv(args.out_all, sep=";", index=False)
    print(f"[wrote] {args.out_all}  ({len(out)} rows, clock-{args.clock_offset_h}h -> UTC)")

    # clear-sky day classification (pvlib Ineichen reference)
    loc = pvlib.location.Location(args.lat, args.lon, altitude=args.alt, tz="UTC")
    df["date"] = df["timestamp_utc"].dt.strftime("%Y-%m-%d")
    clear = []
    for day, d in df.groupby("date"):
        t = pd.DatetimeIndex(d["timestamp_utc"])
        cs = loc.get_clearsky(t, model="ineichen")["ghi"].to_numpy()
        meas = d["ghi"].to_numpy()
        m = cs > 20
        if m.sum() < 20:
            continue
        ratio = float(meas[m].sum() / cs[m].sum())
        if args.clear_min <= ratio <= args.clear_max:
            clear.append((day, round(ratio, 3)))
    cl = pd.DataFrame(clear, columns=["date", "clearness_ratio"]).sort_values("date")
    cl.to_csv(args.out_clear, index=False)
    print(f"[wrote] {args.out_clear}  ({len(cl)} clear-sky days of "
          f"{df['date'].nunique()} total = {len(cl)/df['date'].nunique()*100:.0f}%)")
    if len(cl):
        by_year = cl["date"].str[:4].value_counts().sort_index()
        print("  clear days per year:", dict(by_year))


if __name__ == "__main__":
    main()
