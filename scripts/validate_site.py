r"""Generalised pyranometer validation for ANY site (Thissio, Thessaloniki, ...).

A site-agnostic version of validate_thissio.py. Reads the in-situ pyranometer from
either a tidy UTC CSV (e.g. the AUTh-LAP Thessaloniki record from
read_pyranometer_lap.py: timestamp_utc;sza_deg;ghi_wm2;...) or the Thissio XLSX
(TOTAL/DIFFUSE), aligns to CAMS McClear, optionally to the twin lux, applies a
clear-sky filter, and reports the arms:

  cams_baseline        CAMS McClear vs pyranometer (always; the bar to beat)
  global_transfer      pooled A-J LINEAR fit, applied to this site unseen      (needs --lux + --pooled)
  global_transfer_phys pooled A-J PHYSICAL fit (Erbs kd), applied unseen       (needs --lux + --pooled)
  thissio_heldout*     k-fold linear / physical on this site                  (needs --lux)

Clock: pyranometer timestamps are made UTC via --clock-offset-h (0 for the LAP
'TIME_UT' record; 2 for the Thissio EET clock). Diffuse fraction kd: MEASURED if
the pyranometer has a diffuse channel (Thissio), else MODELLED with Erbs (1982)
from GHI + geometry (Thessaloniki is global-only). Clear-sky selection: measured
kd<=--kd-max where available, else restrict to --clear-days (the classifier output).

Usage (Thessaloniki, CAMS-only until twin lux exists):
  uv run python scripts/validate_site.py --site Thessaloniki \
    --pyrano-csv  ".../pyranometer_thessaloniki_utc.csv" \
    --cams        ".../raw_GHI/Location_LAP_Thessaloniki.csv" \
    --clear-days  ".../pyranometer_thessaloniki_clearsky_days.csv" \
    --clock-offset-h 0 --lat 40.6334 --lon 22.9570 --alt 60 \
    --out ".../results/thessaloniki_validation_summary.csv"
  # add  --lux ".../lux_csv/lux_Location_Thessaloniki.csv" --pooled ".../lux_ghi_monolithic.csv"
  # once the Thessaloniki twin renders exist.
"""
from __future__ import annotations

import argparse
import datetime as _dt
import os

import numpy as np
import pandas as pd


def _metrics(meas, pred):
    meas = np.asarray(meas, float); pred = np.asarray(pred, float)
    k = np.isfinite(meas) & np.isfinite(pred); meas, pred = meas[k], pred[k]
    if len(meas) < 3:
        return {}
    e = pred - meas; rmse = float(np.sqrt(np.mean(e ** 2)))
    ss = float(np.sum(e ** 2)); tot = float(np.sum((meas - meas.mean()) ** 2))
    return dict(n=int(len(meas)), rmse=rmse, mbe=float(e.mean()),
                r2=1 - ss / tot if tot > 0 else float("nan"),
                nrmse=rmse / meas.mean() * 100 if meas.mean() else float("nan"))


def _read_pyrano(args):
    """Return t(UTC), meas(GHI), [kd_meas], [sza_deg]. CSV or Thissio XLSX."""
    if args.pyrano_csv:
        d = pd.read_csv(args.pyrano_csv)
        d["t"] = pd.to_datetime(d["timestamp_utc"], utc=True, format="ISO8601") \
            - pd.Timedelta(hours=args.clock_offset_h)
        d["meas"] = pd.to_numeric(d["ghi_wm2"], errors="coerce")
        d["sza"] = pd.to_numeric(d.get("sza_deg", np.nan), errors="coerce")
        dhi = d["diffuse_wm2"] if "diffuse_wm2" in d.columns else None
        d["kd_meas"] = (pd.to_numeric(dhi, errors="coerce") / d["meas"].replace(0, np.nan)
                        if dhi is not None else np.nan)
        return d[["t", "meas", "kd_meas", "sza"]].dropna(subset=["t", "meas"])
    # Thissio XLSX (TOTAL/DIFFUSE)
    import openpyxl
    wb = openpyxl.load_workbook(args.pyrano_xlsx, read_only=True, data_only=True)
    rows = [r[:7] for sn in wb.sheetnames for r in wb[sn].iter_rows(min_row=2, values_only=True)
            if r[0] is not None]
    df = pd.DataFrame(rows, columns=["Y", "M", "D", "h", "mi", "tot", "dif"])
    df["t"] = (pd.to_datetime(dict(year=df.Y, month=df.M, day=df.D, hour=df.h, minute=df.mi))
               - pd.Timedelta(hours=args.clock_offset_h)).dt.tz_localize("UTC")
    df["meas"] = pd.to_numeric(df["tot"], errors="coerce")
    df["kd_meas"] = pd.to_numeric(df["dif"], errors="coerce") / df["meas"].replace(0, np.nan)
    df["sza"] = np.nan
    return df[["t", "meas", "kd_meas", "sza"]].dropna(subset=["t", "meas"])


def _kfold(lux, meas, build, k=5, seed=0):
    rng = np.random.default_rng(seed); idx = rng.permutation(len(lux))
    pred = np.full(len(lux), np.nan)
    for f in np.array_split(idx, k):
        tr = np.setdiff1d(idx, f); pred[f] = build(tr, f)
    return pred


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--site", default="site")
    p.add_argument("--pyrano-csv"); p.add_argument("--pyrano-xlsx")
    p.add_argument("--cams", required=True)
    p.add_argument("--lux"); p.add_argument("--pooled")
    p.add_argument("--clear-days", help="classifier CSV (date,clear) for sites without measured diffuse")
    p.add_argument("--lat", type=float, required=True)
    p.add_argument("--lon", type=float, required=True)
    p.add_argument("--alt", type=float, default=0.0)
    p.add_argument("--clock-offset-h", type=float, default=0.0)
    p.add_argument("--kd-max", type=float, default=0.25)
    p.add_argument("--ghi-min", type=float, default=50.0)
    p.add_argument("--elev-min", type=float, default=5.0)
    p.add_argument("--out", required=True); p.add_argument("--scatter")
    args = p.parse_args()
    if not (args.pyrano_csv or args.pyrano_xlsx):
        p.error("need --pyrano-csv or --pyrano-xlsx")

    import pvlib
    from solar_twin import io_cams, physical_model as pm

    pyr = _read_pyrano(args)
    cdf, _ = io_cams.read_cams(args.cams)
    ci = io_cams.to_instantaneous(cdf)
    cams = ci[["mid_utc", "ghi_wm2"]].rename(columns={"mid_utc": "t", "ghi_wm2": "cams"})

    j = pd.merge_asof(pyr.sort_values("t"), cams.sort_values("t"), on="t",
                      direction="nearest", tolerance=pd.Timedelta(minutes=8)).dropna(subset=["cams"])
    # elevation: from measured SZA if present, else pvlib
    if j["sza"].notna().any():
        j["elev"] = 90.0 - j["sza"]
    else:
        sp = pvlib.solarposition.spa_python(pd.DatetimeIndex(j["t"]), args.lat, args.lon, altitude=args.alt)
        j["elev"] = sp["apparent_elevation"].values
    j["doy"] = j["t"].dt.dayofyear
    j["kd"] = np.where(j["kd_meas"].notna(), j["kd_meas"],
                       pm.erbs_kd(j["meas"].to_numpy(), j["elev"].to_numpy(), j["doy"].to_numpy()))
    j = j[(j.elev > args.elev_min) & (j.cams > 20)].copy()

    # ---- clear-sky filter ----
    n_all = len(j)
    j = j[(j.meas > args.ghi_min)].copy()
    if j["kd_meas"].notna().any():
        j = j[j.kd <= args.kd_max].copy(); filt = f"measured kd<={args.kd_max}"
    elif args.clear_days:
        cd = pd.read_csv(args.clear_days)
        clear = set(pd.to_datetime(cd[cd["clear"].astype(bool)]["date"]).dt.date)
        j = j[j["t"].dt.date.isin(clear)].copy(); filt = f"clear-day classifier ({len(clear)} days)"
    else:
        filt = "GHI>min only (no diffuse, no --clear-days)"
    print(f"[{args.site}] {n_all} matched -> {len(j)} kept [{filt}]")
    if len(j) < 10:
        print("[abort] too few rows"); return

    meas = j.meas.to_numpy()
    arms = {"cams_baseline": j["cams"].to_numpy()}

    if args.lux:
        lux = pd.read_csv(args.lux, sep=";")
        tcol = "timestamp_utc" if "timestamp_utc" in lux.columns else lux.columns[0]
        lux["t"] = pd.to_datetime(lux[tcol], utc=True); lux["lux"] = pd.to_numeric(lux["lux"], errors="coerce")
        j = pd.merge_asof(j.sort_values("t"), lux[["t", "lux"]].sort_values("t"), on="t",
                          direction="nearest", tolerance=pd.Timedelta(minutes=8)).dropna(subset=["lux"])
        meas = j.meas.to_numpy(); luxv = j.lux.to_numpy()
        arms["cams_baseline"] = j["cams"].to_numpy()
        # linear k-fold
        arms["heldout_linear"] = _kfold(luxv, meas,
            lambda tr, f: np.polyval(np.polyfit(luxv[tr], meas[tr], 1), luxv[f]))
        # physical k-fold (horizontal-open, kd as computed)
        sp = pvlib.solarposition.spa_python(pd.DatetimeIndex(j["t"]), args.lat, args.lon, altitude=args.alt)
        tdf = pd.DataFrame(dict(location_id=args.site, lux=luxv, ghi=meas,
                                solar_elevation_deg=j.elev.to_numpy(), solar_azimuth_deg=sp["azimuth"].values,
                                air_mass=np.nan_to_num(pvlib.atmosphere.get_relative_airmass(
                                    90 - j.elev.to_numpy(), model="kastenyoung1989"), nan=40.0),
                                kd=j.kd.to_numpy()))

        def _bp(tr, f):
            m = pm.fit_physical(tdf.iloc[tr], {}, "split"); return pm.predict_physical(m, tdf.iloc[f], {})
        arms["heldout_physical"] = _kfold(luxv, meas, _bp)
        if args.pooled and os.path.exists(args.pooled):
            pool = pd.read_csv(args.pooled, sep=";" if open(args.pooled).readline().count(";") else ",")
            pool = pool.dropna(subset=["lux", "ghi"])
            ga, gb = np.polyfit(pool.lux.astype(float), pool.ghi.astype(float), 1)
            arms["global_transfer"] = ga * luxv + gb
            need = {"solar_elevation_deg", "solar_azimuth_deg", "air_mass", "day_of_year"}
            if need.issubset(pool.columns):
                pl = pool.dropna(subset=list(need)).copy()
                pl["kd"] = pm.erbs_kd(pl.ghi, pl.solar_elevation_deg, pl.day_of_year)
                mph = pm.fit_physical(pl, {}, "split")
                arms["global_transfer_phys"] = pm.predict_physical(mph, tdf, {})

    print(f"{'arm':>20} | {'n':>5} {'RMSE':>6} {'MBE':>6} {'R2':>6} {'nRMSE%':>6}")
    rows = []
    for name, pred in arms.items():
        m = _metrics(meas, pred)
        if m:
            rows.append(("all", name, m))
            print(f"{name:>20} | {m['n']:>5} {m['rmse']:>6.1f} {m['mbe']:>+6.1f} {m['r2']:>6.3f} {m['nrmse']:>6.1f}")

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    cols = ["n", "rmse", "mbe", "r2", "nrmse"]
    with open(args.out, "w", encoding="utf-8") as fh:
        fh.write(f"# site={args.site} generated={_dt.date.today()} filter={filt} "
                 f"clock_offset_h={args.clock_offset_h}\n")
        fh.write("stratum,arm," + ",".join(cols) + "\n")
        for st, name, m in rows:
            fh.write(f"{st},{name}," + ",".join(f"{m[c]:.4f}" for c in cols) + "\n")
    print(f"[wrote] {args.out}")

    if args.scatter and "heldout_physical" in arms:
        try:
            import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
            fig, ax = plt.subplots(figsize=(6, 6))
            ax.scatter(meas, arms["heldout_physical"], s=8, alpha=0.4, color="#1d9e75", label="twin physical")
            ax.scatter(meas, arms["cams_baseline"], s=8, alpha=0.4, color="#d85a30", marker="^", label="CAMS")
            lim = float(np.nanmax([meas.max(), np.nanmax(arms["cams_baseline"])])) * 1.05
            ax.plot([0, lim], [0, lim], "k--", lw=1)
            ax.set_xlabel("measured GHI (W/m²)"); ax.set_ylabel("predicted GHI (W/m²)")
            ax.set_title(f"{args.site}: twin vs CAMS, clear-sky"); ax.legend()
            fig.tight_layout(); fig.savefig(args.scatter, dpi=150)
            print(f"[wrote] {args.scatter}")
        except Exception as e:  # noqa: BLE001
            print(f"[warn] scatter skipped: {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
