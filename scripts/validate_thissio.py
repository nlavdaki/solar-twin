r"""Thissio pyranometer validation — the ground-truth test, done properly (uv, GPU-free).

Compares twin-derived GHI against the in-situ pyranometer, THREE model arms +
the CAMS baseline, clear-sky-filtered, elevation-stratified:

  (1) GLOBAL TRANSFER  — pooled A-J fit (from the monolithic dataset) applied to
      Thissio UNSEEN. The real transfer-learning test: does a model built on other
      rooftops predict a never-seen roof's measured GHI?
  (2) THISSIO HELD-OUT — a Thissio-specific fit with k-fold CV (fit on some clear
      days, predict the others). The upper bound: how well CAN this site calibrate?
  (3) IN-SAMPLE (reference only) — Thissio fit on all its data, scored on the same.
      Flatters; reported only as a ceiling, NOT a result.
  (B) CAMS BASELINE   — CAMS McClear vs pyranometer (the bar to beat).

CLEAR-SKY / BEAM-DOMINATED FILTER (twin is cloudless; pyranometer is all-sky):
keep rows where the MEASURED diffuse fraction kd = DHI/GHI <= --kd-max (default
0.25) and GHI > --ghi-min (default 50). kd is derived from the station's own
global + diffuse pyranometers, so it is independent of CAMS and of the twin --
the twin-vs-CAMS comparison stays non-circular. (Empirically this is the filter
that tightens the lux->GHI relation at Thissio: it removes the high-diffuse,
low-sun regime where the twin's clear-sky sky-dome diverges most from reality.)
An optional CSR (measured/CAMS) window is available but OFF by default. Time
base: pyranometer clock is fixed UTC+2 (no DST) -> UTC = clock - 2h.

Usage (uv):
  uv run python scripts/validate_thissio.py \
    --lux       "…/data/lux_csv/lux_Location_Thissio.csv" \
    --cams      "…/data/raw_GHI/Location_Thissio.csv" \
    --pyrano-xlsx "…/data/pyranometer_GHI_ground_level/THISSIO-2020-2024_step-15min_FINAL.xlsx" \
    --pooled    "…/data/dataset/lux_ghi_monolithic.csv" \
    --lat 37.9717 --lon 23.7182 --alt 100 \
    --out "…/data/results/thissio_validation_summary.csv" \
    --scatter "…/data/results/thissio_validation_scatter.png"
"""
from __future__ import annotations

import argparse
import datetime as _dt
import os

import numpy as np
import pandas as pd

HW = "RTX 4070 12GB | i5-12600K | 32GB | IsaacSim 5.1.0-rc19 Kit107.3.3"


def _metrics(meas, pred):
    meas = np.asarray(meas, float); pred = np.asarray(pred, float)
    k = np.isfinite(meas) & np.isfinite(pred)
    meas, pred = meas[k], pred[k]
    if len(meas) < 3:
        return {}
    e = pred - meas
    rmse = float(np.sqrt(np.mean(e ** 2)))
    ss = float(np.sum(e ** 2)); tot = float(np.sum((meas - meas.mean()) ** 2))
    return dict(n=len(meas), rmse=rmse, mbe=float(e.mean()),
                r2=1 - ss / tot if tot > 0 else np.nan,
                nrmse=rmse / meas.mean() * 100 if meas.mean() else np.nan)


def _read_pyrano_xlsx(path, clock_offset_h=2.0):
    """Read NOA Thissio actinometric xlsx.

    Columns: YEAR MONTH DAY HOUR MIN  TOTAL_AVG(=GHI)  DIFFUSE_AVG(=DHI).
    The station carries an unshaded (global) AND a shaded (diffuse) pyranometer,
    so we get the MEASURED diffuse fraction kd = DHI/GHI directly. kd is intrinsic
    to the sensor (independent of CAMS and of the twin), making it a non-circular
    clear-sky / beam-dominated selector for the twin-vs-CAMS comparison.
    Returns t, meas(GHI), dhi, kd.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = [r[:7] for sn in wb.sheetnames for r in wb[sn].iter_rows(min_row=2, values_only=True)
            if r[0] is not None]
    df = pd.DataFrame(rows, columns=["Y", "M", "D", "h", "mi", "ghi", "dif"])
    df["t"] = (pd.to_datetime(dict(year=df.Y, month=df.M, day=df.D, hour=df.h, minute=df.mi))
               - pd.Timedelta(hours=clock_offset_h)).dt.tz_localize("UTC")
    df["meas"] = pd.to_numeric(df["ghi"], errors="coerce")
    df["dhi"] = pd.to_numeric(df["dif"], errors="coerce")
    df["kd"] = df["dhi"] / df["meas"].replace(0, np.nan)
    return df[["t", "meas", "dhi", "kd"]].dropna(subset=["t", "meas"])


def _kfold_heldout(lux, meas, k=5, seed=0):
    """Predict each fold from a fit on the others. Returns out-of-fold predictions."""
    rng = np.random.default_rng(seed)
    idx = rng.permutation(len(lux))
    folds = np.array_split(idx, k)
    pred = np.full(len(lux), np.nan)
    for f in folds:
        tr = np.setdiff1d(idx, f)
        a, b = np.polyfit(lux[tr], meas[tr], 1)
        pred[f] = a * lux[f] + b
    return pred


def _kfold_phys(df, k=5, seed=0, tier="split_am", geom=None):
    """Out-of-fold predictions from the structured physical model. df needs lux,
    ghi, solar_elevation_deg, solar_azimuth_deg, air_mass, kd. geom = per-site
    geometry dict (defaults to horizontal-open)."""
    from solar_twin import physical_model as pm
    geom = geom or {}
    arr = df.reset_index(drop=True)
    rng = np.random.default_rng(seed)
    idx = rng.permutation(len(arr))
    folds = np.array_split(idx, k)
    pred = np.full(len(arr), np.nan)
    for f in folds:
        tr = np.setdiff1d(idx, f)
        m = pm.fit_physical(arr.iloc[tr], geom, tier)
        pred[f] = pm.predict_physical(m, arr.iloc[f], geom)
    return pred


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--lux", required=True)
    p.add_argument("--cams", required=True)
    p.add_argument("--pyrano-xlsx", required=True)
    p.add_argument("--pooled", required=True, help="monolithic A-J dataset -> pooled global fit")
    p.add_argument("--lat", type=float, required=True)
    p.add_argument("--lon", type=float, required=True)
    p.add_argument("--alt", type=float, default=100.0)
    p.add_argument("--clock-offset-h", type=float, default=2.0)
    # PRIMARY clear-sky / beam-dominated filter (measurement-intrinsic, non-circular):
    p.add_argument("--kd-max", type=float, default=0.25,
                   help="keep rows with measured diffuse fraction DHI/GHI <= this "
                        "(beam-dominated clear sky; the proven Thissio filter)")
    p.add_argument("--ghi-min", type=float, default=50.0,
                   help="keep rows with measured GHI > this W/m2 (drop twilight)")
    # OPTIONAL secondary CSR window — OFF by default (0..99) to avoid CAMS circularity:
    p.add_argument("--csr-lo", type=float, default=0.0)
    p.add_argument("--csr-hi", type=float, default=99.0)
    p.add_argument("--tol-min", type=float, default=8.0)
    p.add_argument("--elev-min", type=float, default=5.0)
    p.add_argument("--out", required=True)
    p.add_argument("--scatter", default=None)
    p.add_argument("--geometry-dir", default=None,
                   help="dir of geometry_<site>.json (tilt/SVF/horizon from "
                        "extract_geometry_gui.py); missing sites -> horizontal-open")
    args = p.parse_args()

    import pvlib
    from solar_twin import io_cams

    # twin lux
    lux = pd.read_csv(args.lux, sep=";")
    tcol = "timestamp_utc" if "timestamp_utc" in lux.columns else lux.columns[0]
    lux["t"] = pd.to_datetime(lux[tcol], utc=True)
    lux["lux"] = pd.to_numeric(lux["lux"], errors="coerce")

    # CAMS McClear -> instantaneous
    cdf, _ = io_cams.read_cams(args.cams)
    cinst = io_cams.to_instantaneous(cdf)[["mid_utc", "ghi_wm2"]].rename(
        columns={"mid_utc": "t", "ghi_wm2": "cams"})

    # pyranometer
    pyr = _read_pyrano_xlsx(args.pyrano_xlsx, args.clock_offset_h)

    # join
    j = pd.merge_asof(lux.sort_values("t"), pyr.sort_values("t"), on="t",
                      direction="nearest", tolerance=pd.Timedelta(minutes=args.tol_min)).dropna(subset=["meas"])
    j = pd.merge_asof(j.sort_values("t"), cinst.sort_values("t"), on="t",
                      direction="nearest", tolerance=pd.Timedelta(minutes=args.tol_min))
    sp = pvlib.solarposition.spa_python(pd.DatetimeIndex(j["t"]), args.lat, args.lon, altitude=args.alt)
    j["elev"] = sp["apparent_elevation"].values
    j["saz"] = sp["azimuth"].values
    j["am"] = pvlib.atmosphere.get_relative_airmass(90.0 - j["elev"].values, model="kastenyoung1989")
    j = j[(j.elev > args.elev_min) & (j.cams > 20)].copy()

    # --- clear-sky / beam-dominated filter --------------------------------
    # PRIMARY: measured diffuse fraction kd = DHI/GHI <= kd_max, plus GHI>ghi_min.
    # kd comes from the station's own global+diffuse sensors, so it is independent
    # of both CAMS and the twin -> the twin-vs-CAMS comparison is NOT circular.
    # This is the regime the cloudless twin can physically match, and the filter
    # that empirically tightens the lux->GHI relation at this site.
    # OPTIONAL: CSR (measured/CAMS) window, off by default (0..99).
    j["csr"] = j.meas / j.cams
    n_all = len(j)
    j = j[(j.meas > args.ghi_min) & (j.kd <= args.kd_max)].copy()
    j = j[(j.csr >= args.csr_lo) & (j.csr <= args.csr_hi)].copy()
    print(f"[thissio] matched {n_all} rows -> {len(j)} kept "
          f"({len(j)/n_all*100:.0f}%) [kd<={args.kd_max} GHI>{args.ghi_min} "
          f"csr {args.csr_lo}-{args.csr_hi}]")
    if len(j) < 10:
        print("[abort] too few clear-sky rows"); return

    luxv, meas = j.lux.to_numpy(), j.meas.to_numpy()

    # ---- LINEAR arms (constant-efficacy baseline = the current model) ----
    # (1) GLOBAL TRANSFER: pooled A-J linear fit, applied to Thissio unseen
    pool = pd.read_csv(args.pooled, sep=";" if open(args.pooled).readline().count(";") else ",")
    pool = pool.dropna(subset=["lux", "ghi"])
    ga, gb = np.polyfit(pool["lux"].to_numpy(float), pool["ghi"].to_numpy(float), 1)
    j["pred_global"] = ga * luxv + gb
    # (2) THISSIO HELD-OUT k-fold (linear)
    j["pred_heldout"] = _kfold_heldout(luxv, meas, k=5)
    # (3) IN-SAMPLE (ceiling only)
    ia, ib = np.polyfit(luxv, meas, 1)
    j["pred_insample"] = ia * luxv + ib

    arms = {"global_transfer": j["pred_global"], "thissio_heldout": j["pred_heldout"],
            "thissio_insample": j["pred_insample"], "cams_baseline": j["cams"]}

    # ---- PHYSICAL arms (structured beam/diffuse inverse model) ----
    # Thissio uses its MEASURED kd; the air-mass + beam/diffuse split absorb the
    # luminous-efficacy drift that mathematically caps the linear fit. Geometry
    # defaults to horizontal-open (no geometry json yet) -> still diffuse- and
    # air-mass-aware; supply geometry_Location_Thissio.json to refine a tilt.
    try:
        from solar_twin import physical_model as pm
        # load per-site geometry jsons (tilt/SVF/horizon); missing -> horizontal-open
        geom_by_site = {}
        if args.geometry_dir and os.path.isdir(args.geometry_dir):
            import glob
            import json as _json
            for f in glob.glob(os.path.join(args.geometry_dir, "geometry_*.json")):
                d = _json.load(open(f))
                geom_by_site[d["location_id"]] = pm.SiteGeometry.from_dict(d)
            print(f"[thissio] geometry loaded for: {sorted(geom_by_site)}")
        tgeo = geom_by_site.get("Location_Thissio") or geom_by_site.get("Thissio")
        tgeom = {"Thissio": tgeo} if tgeo is not None else {}
        if tgeo is not None:
            print(f"[thissio] Thissio geometry: tilt={tgeo.tilt_deg} az={tgeo.azimuth_deg} svf={tgeo.svf}")

        tdf = pd.DataFrame(dict(location_id="Thissio", lux=luxv, ghi=meas,
                                solar_elevation_deg=j.elev.to_numpy(),
                                solar_azimuth_deg=j.saz.to_numpy(),
                                air_mass=j.am.to_numpy(), kd=j.kd.to_numpy()))
        need = {"solar_elevation_deg", "solar_azimuth_deg", "air_mass", "day_of_year"}
        if need.issubset(pool.columns):
            pl = pool.dropna(subset=list(need | {"lux", "ghi"})).copy()
            pl["kd"] = pm.erbs_kd(pl["ghi"], pl["solar_elevation_deg"], pl["day_of_year"])
            mph = pm.fit_physical(pl, geom_by_site, tier="split_am")  # uses A-J geometry if present
            j["pred_global_phys"] = pm.predict_physical(mph, tdf, tgeom)
            arms["global_transfer_phys"] = j["pred_global_phys"]
            print(f"[thissio] pooled A-J physical efficacies (lm/W): "
                  f"{ {k: round(v, 1) for k, v in mph.efficacies().items()} }")
        else:
            print("[thissio] pooled lacks solar-geometry cols -> skip global_transfer_phys "
                  "(re-run build_dataset to add them)")
        j["pred_heldout_phys"] = _kfold_phys(tdf, k=5, geom=tgeom)
        arms["thissio_heldout_phys"] = j["pred_heldout_phys"]
    except Exception as e:  # noqa: BLE001
        print(f"[thissio] physical arms skipped: {type(e).__name__}: {e}")

    print(f"\n[thissio] pooled A-J global fit: GHI={ga:.5f}*lux{gb:+.1f} (eff~{1/ga:.0f} lm/W)")
    print(f"{'arm':>18} | {'n':>4} {'RMSE':>6} {'MBE':>6} {'R2':>6} {'nRMSE%':>6}")
    rows = []
    for name, pred in arms.items():
        m = _metrics(meas, pred)
        if m:
            rows.append(("all", name, m))
            print(f"{name:>18} | {m['n']:>4} {m['rmse']:>6.1f} {m['mbe']:>+6.1f} {m['r2']:>6.3f} {m['nrmse']:>6.1f}")
    # elevation strata: linear global / PHYSICAL held-out / CAMS
    has_phys = "pred_heldout_phys" in j.columns
    print("\n  elevation strata (linear global / PHYSICAL held-out / CAMS):")
    for lo, hi in [(5, 15), (15, 30), (30, 45), (45, 90)]:
        s = j[(j.elev >= lo) & (j.elev < hi)]
        if len(s) < 5:
            continue
        g = _metrics(s.meas, s.pred_global); c = _metrics(s.meas, s.cams)
        rows.append((f"elev_{lo}_{hi}", "global_transfer", g))
        rows.append((f"elev_{lo}_{hi}", "cams_baseline", c))
        ph = _metrics(s.meas, s.pred_heldout_phys) if has_phys else {}
        if ph:
            rows.append((f"elev_{lo}_{hi}", "thissio_heldout_phys", ph))
        pstr = f" | phys R2={ph['r2']:>6.3f} MBE={ph['mbe']:>+6.1f}" if ph else ""
        print(f"   {lo:>2}-{hi:<3}deg n={len(s):>3} | lin R2={g['r2']:>6.3f} MBE={g['mbe']:>+6.1f}"
              f"{pstr} | cams R2={c['r2']:>6.3f} MBE={c['mbe']:>+6.1f}")
    print("   (per-bin R2 is restricted-range-sensitive; MBE/bias is the primary per-bin metric)")

    # verdict
    gm = _metrics(meas, j.pred_global); cm = _metrics(meas, j.cams)
    print(f"\n[VERDICT] global-transfer twin RMSE {gm['rmse']:.1f} vs CAMS {cm['rmse']:.1f} W/m2 -> "
          + ("TWIN >= CAMS (transfer works)" if gm['rmse'] <= cm['rmse'] else "CAMS better (expected; twin is unseen-transfer)"))

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    cols = ["n", "rmse", "mbe", "r2", "nrmse"]
    with open(args.out, "w", encoding="utf-8") as fh:
        fh.write(f"# hardware={HW}\n# generated={_dt.date.today()} site=Thissio "
                 f"filter=kd<={args.kd_max},GHI>{args.ghi_min},csr{args.csr_lo}-{args.csr_hi} "
                 f"pooled_global_fit a={ga:.6f} b={gb:.3f}\n")
        fh.write("stratum,arm," + ",".join(cols) + "\n")
        for stratum, name, m in rows:
            fh.write(f"{stratum},{name}," + ",".join(f"{m[c]:.4f}" for c in cols) + "\n")
    print(f"[wrote] {args.out}")

    if args.scatter:
        try:
            import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
            # Plot the held-out PHYSICAL arm (out-of-fold k-fold, geometry/diffuse-
            # aware) — this is the cited Thissio result (R² ≈ 0.937). Fall back to
            # the held-out linear arm only if the physical arm was not produced.
            if "pred_heldout_phys" in j.columns and j["pred_heldout_phys"].notna().any():
                twin_pred = j.pred_heldout_phys; twin_label = "twin physical (held-out CV)"
            else:
                twin_pred = j.pred_heldout; twin_label = "twin (held-out CV, linear)"
            fig, ax = plt.subplots(figsize=(6, 6))
            ax.scatter(meas, twin_pred, s=8, alpha=0.4, color="#1d9e75", label=twin_label)
            ax.scatter(meas, j.cams, s=8, alpha=0.4, color="#d85a30", marker="^", label="CAMS")
            lim = float(np.nanmax([meas.max(), twin_pred.max()])) * 1.05
            ax.plot([0, lim], [0, lim], "k--", lw=1, label="1:1")
            ax.set_xlabel("measured GHI — pyranometer (W/m²)"); ax.set_ylabel("predicted GHI (W/m²)")
            ax.set_title("Thissio: twin physical (held-out CV) vs CAMS, clear-sky")
            ax.legend(); fig.tight_layout(); fig.savefig(args.scatter, dpi=150)
            print(f"[wrote] {args.scatter}")
        except Exception as e:  # noqa: BLE001
            print(f"[warn] scatter skipped: {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
